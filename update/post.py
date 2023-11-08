import re
import numpy as np
from openpyxl.workbook import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

def post_processing(results, savePath):
    '''
    This function makes xlsx files from Boxes and Texts.
    1. Loop Boxes and Texts according to document.
    2. In a document, loop boxes and texts according to page.
    3. SR number modification
    4. Consider cell swrap, thin
    5. Save
    '''
    filenames = list(results.keys())
    if len(filenames) > 0:

        col_title = list(results[filenames[0]].keys())
        pre_rows = 0 # considering multi tables in a page.
        wb = Workbook()
        ws = wb.active
        ws.title = "new table"
        ws.cell(pre_rows+1,1).value = 'File Name'
        ws.cell(pre_rows+1,1).font = Font(bold=True)            
        for i in range(len(col_title)):
            ws.cell(pre_rows+1,i+2).value = col_title[i]
            ws.cell(pre_rows+1,i+2).font = Font(bold=True)    
        pre_rows += 1
        thin_border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')) 
        ### loopping every document in multi document ###
        for kk, filename in enumerate(filenames):
            result = results[filename]
            ws.cell(pre_rows+1,1).value = filename
            for i, val in enumerate(result.values()):
                ws.cell(pre_rows+1,i+2).value = val
            pre_rows += 1
            
        # cell swrap, thin
        row_no = 1
        for i in ws.rows:
            for j in range(len(i)):
                ws[get_column_letter(j+1)+str(row_no)].alignment = Alignment(wrap_text=True, vertical='center',horizontal='center')
                ws.cell(row=row_no, column=j + 1).border = thin_border
            row_no = row_no + 1  
        # column width
        column_width = [20] * (len(col_title)+1)
        column_width[4] = 40
        for i in range(len(col_title)):
            ws.column_dimensions[get_column_letter(i+1)].width = column_width[i]
        ws.sheet_view.zoomScale = 85
        # save
        wb.save(savePath)
    else:
        print("=== No results ===")

    return None